#!/usr/bin/env python3
"""Exporta hojas de catálogo del Excel técnico IRIS a JSON para la app.

Uso:
  python3 tools/export_catalog_excel_to_json.py \\
    docs/catalog/lista_de_material_tecnico_v1_7.xlsx \\
    docs/catalog/

Escribe:
  cameras_modos_v1_7.json
  lenses_v1_7.json
  lights_v1_7.json

Requiere openpyxl. El paquete Dart `excel` no abre bien libros ricos.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as e:  # pragma: no cover
    raise SystemExit("Instala openpyxl: pip install openpyxl") from e


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"-?\d+(\.\d+)?", str(v).replace(",", "."))
    return float(m.group(0)) if m else None


def _to_int(v):
    f = _to_float(v)
    return int(round(f)) if f is not None else None


def _truthy(v) -> bool:
    return v in (1, True, "1", "true", "True", "sí", "si", "yes")


def _sheet_rows(wb, name: str):
    rows = list(wb[name].iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return headers, rows[1:]


def export_cameras(wb, dst: Path) -> None:
    cam_h, cam_rows = _sheet_rows(wb, "Cámaras")
    mode_h, mode_rows = _sheet_rows(wb, "Modos_Cámara")

    modes_by: dict[str, list] = {}
    skipped_modes = 0
    for row in mode_rows:
        d = {mode_h[i]: (row[i] if i < len(row) else None) for i in range(len(mode_h))}
        cid = d.get("camera_external_id")
        if not cid:
            continue
        name = d.get("modo_sensor")
        gw = _to_float(d.get("gate_ancho_mm"))
        gh = _to_float(d.get("gate_alto_mm"))
        if not name or gw is None or gh is None:
            skipped_modes += 1
            continue
        mode = {
            "name": str(name),
            "widthMm": gw,
            "heightMm": gh,
            "cropFactor": _to_float(d.get("crop_x")) or 1.0,
        }
        cy = _to_float(d.get("crop_y"))
        if cy is not None:
            mode["cropY"] = cy
        rx, ry = _to_int(d.get("res_x")), _to_int(d.get("res_y"))
        if rx is not None:
            mode["maxWidthPx"] = rx
        if ry is not None:
            mode["maxHeightPx"] = ry
        if d.get("aspect_ratio"):
            mode["aspectRatio"] = str(d["aspect_ratio"])
        if d.get("codec"):
            mode["codec"] = str(d["codec"])
        fm = _to_float(d.get("fps_max"))
        if fm is not None:
            mode["fpsMax"] = fm
        modes_by.setdefault(str(cid), []).append(mode)

    cameras = []
    for row in cam_rows:
        d = {cam_h[i]: (row[i] if i < len(row) else None) for i in range(len(cam_h))}
        eid = d.get("external_id")
        if not eid or _truthy(d.get("es_custom")):
            continue
        brand, model = d.get("marca"), d.get("modelo")
        w, h = _to_float(d.get("sensor_ancho_mm")), _to_float(d.get("sensor_alto_mm"))
        if not brand or not model or w is None or h is None:
            continue
        entry = {
            "externalId": str(eid),
            "brand": str(brand),
            "model": str(model),
            "sensorWidthMm": w,
            "sensorHeightMm": h,
            "vintage": _truthy(d.get("vintage")),
            "lukaCompatible": _truthy(d.get("luka_compatible")),
            "sensorModes": modes_by.get(str(eid), []),
        }
        if d.get("mount"):
            entry["mountType"] = str(d["mount"])
        ni = _to_int(d.get("iso_base")) or _to_int(d.get("iso_base_1"))
        if ni is not None:
            entry["nativeIso"] = ni
        dr = _to_float(d.get("rango_dinamico_stops"))
        if dr is not None:
            entry["dynamicRangeStops"] = dr
        if d.get("tipo_sensor"):
            entry["colorScience"] = str(d["tipo_sensor"])
        if d.get("fuente_datos"):
            entry["manufacturerUrl"] = str(d["fuente_datos"])
        cameras.append(entry)

    out = dst / "cameras_modos_v1_7.json"
    out.write_text(json.dumps(cameras, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        f"OK cámaras {len(cameras)} "
        f"({sum(1 for c in cameras if c['sensorModes'])} con modos, "
        f"{sum(len(c['sensorModes']) for c in cameras)} modos; "
        f"{skipped_modes} modos sin gate omitidos) → {out}"
    )


def export_lenses(wb, dst: Path) -> None:
    headers, data = _sheet_rows(wb, "Ópticas")
    lenses = []
    skipped = 0
    for row in data:
        d = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        eid = d.get("external_id")
        if not eid or _truthy(d.get("es_custom")):
            continue
        brand, model = d.get("marca"), d.get("modelo")
        focal = _to_float(d.get("distancia_focal")) or _to_float(d.get("focal_min_mm"))
        tstop = _to_float(d.get("t_stop_min"))
        coverage = d.get("cobertura")
        if not brand or not model or focal is None or tstop is None or not coverage:
            skipped += 1
            continue
        entry = {
            "externalId": str(eid),
            "brand": str(brand),
            "model": str(model),
            "focalLength": focal,
            "minTStop": tstop,
            "formatCoverage": str(coverage),
            "isAnamorphic": _truthy(d.get("anamorfica")),
            "vintage": False,
            "lukaCompatible": False,
        }
        fmin, fmax = _to_float(d.get("focal_min_mm")), _to_float(d.get("focal_max_mm"))
        if fmin is not None:
            entry["focalMin"] = fmin
        if fmax is not None:
            entry["focalMax"] = fmax
        if d.get("mount"):
            entry["mountType"] = str(d["mount"])
        ic = _to_float(d.get("image_circle_mm"))
        if ic is not None:
            entry["imageCircleMm"] = ic
        cf = _to_float(d.get("min_focus_m"))
        if cf is not None:
            entry["closeFocusM"] = cf
        sq = _to_float(d.get("squeeze"))
        if sq is not None:
            entry["squeezeRatio"] = sq
        if d.get("tipo_optica"):
            entry["lensType"] = str(d["tipo_optica"])
        fd = _to_float(d.get("front_diameter_mm"))
        if fd is not None:
            entry["frontDiameterMm"] = fd
        ln = _to_float(d.get("longitud_mm"))
        if ln is not None:
            entry["lengthMm"] = ln
        wg = _to_float(d.get("peso_g"))
        if wg is not None:
            entry["weightKg"] = wg / 1000.0
        lenses.append(entry)

    out = dst / "lenses_v1_7.json"
    out.write_text(json.dumps(lenses, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK ópticas {len(lenses)} (omitidas {skipped}) → {out}")


def export_lights(wb, dst: Path) -> None:
    headers, data = _sheet_rows(wb, "Luces")
    lights = []
    skipped = 0
    for row in data:
        d = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        eid = d.get("external_id")
        if not eid or _truthy(d.get("es_custom")):
            continue
        brand, model = d.get("marca"), d.get("modelo")
        ltype = d.get("tipo_luz") or d.get("tecnologia") or "led_panel"
        power = _to_int(d.get("potencia_w"))
        tmin = _to_int(d.get("temp_min"))
        tmax = _to_int(d.get("temp_max"))
        if (tmin is None or tmax is None) and d.get("cct"):
            nums = re.findall(r"\d+", str(d["cct"]))
            if len(nums) >= 2:
                tmin = tmin or int(nums[0])
                tmax = tmax or int(nums[1])
            elif len(nums) == 1:
                tmin = tmin or int(nums[0])
                tmax = tmax or int(nums[0])
        if not brand or not model or power is None:
            skipped += 1
            continue
        if tmin is None:
            tmin = 3200
        if tmax is None:
            tmax = 5600
        entry = {
            "externalId": str(eid),
            "brand": str(brand),
            "model": str(model),
            "lightType": str(ltype),
            "powerW": power,
            "colorTempMin": tmin,
            "colorTempMax": tmax,
            "lukaCompatible": False,
            "vintage": False,
        }
        cri = _to_int(d.get("cri"))
        if cri is not None:
            entry["cri"] = cri
        tlci = _to_int(d.get("tlci"))
        if tlci is not None:
            entry["tlci"] = tlci
        if d.get("dimming"):
            entry["dimmingType"] = str(d["dimming"])
        lights.append(entry)

    out = dst / "lights_v1_7.json"
    out.write_text(json.dumps(lights, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK luces {len(lights)} (omitidas {skipped}) → {out}")


def export_meta_markdown(wb, dst: Path) -> None:
    """Copia hojas de diseño/meta a markdown (no se importan a Drift)."""
    meta_dir = dst / "design"
    meta_dir.mkdir(parents=True, exist_ok=True)
    sheets = [
        "Mapa_Sistema",
        "Flujo_Eleccion",
        "Control_Datos",
        "Funciones_Futuras",
        "Pendientes_Investigacion",
        "Campos_Tecnicos_Futuros",
        "Fase_2_Ideas",
        "Arquitectura_Dependencias",
        "Verificacion_Fase_1",
        "Leyenda_Verificacion",
    ]
    index_lines = ["# Diseño / meta del Excel v1.7", "", "Hojas de documentación (no Drift).", ""]
    for name in sheets:
        if name not in wb.sheetnames:
            continue
        rows = list(wb[name].iter_rows(values_only=True))
        lines = [f"# {name}", ""]
        for row in rows[:80]:
            cells = ["" if c is None else str(c).replace("\n", " ") for c in row]
            if not any(cells):
                continue
            lines.append("| " + " | ".join(cells[:8]) + " |")
        out = meta_dir / f"{name}.md"
        out.write_text("\n".join(lines) + "\n", encoding="utf-8")
        index_lines.append(f"- [{name}]({name}.md)")
        print(f"OK meta → {out}")
    (meta_dir / "README.md").write_text("\n".join(index_lines) + "\n", encoding="utf-8")


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print(__doc__)
        return 2
    src = Path(argv[1])
    dst = Path(argv[2])
    dst.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(src, data_only=True, read_only=True)
    export_cameras(wb, dst)
    export_lenses(wb, dst)
    export_lights(wb, dst)
    export_meta_markdown(wb, dst)
    wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
